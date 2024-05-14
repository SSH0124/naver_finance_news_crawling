import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

# 테마 페이지에서 테마에 속한 주식명을 가져오는 함수
def get_stocks(theme_url):
    response = requests.get(theme_url)
    soup = BeautifulSoup(response.text, 'html.parser')

    # 각 주식명을 담을 리스트
    stocks = []

    # 테마 페이지에서 각 주식을 나타내는 행을 찾습니다.
    rows = soup.find_all('tr', onmouseover="mouseOver(this)")

    for row in rows:
        # 각 주식명은 <td class="name"> 안에 있습니다.
        stock_name = row.find('td', class_='name').find('a').text.strip()
        stocks.append(stock_name)

    return stocks

# 메인 페이지에서 각 테마의 URL을 가져옵니다.
url = "https://finance.naver.com/sise/theme.naver"
response = requests.get(url)
soup = BeautifulSoup(response.text, 'html.parser')

# 엑셀 파일 생성
wb = Workbook()
ws = wb.active

# 각 테마 페이지에서 주식명을 수집하여 엑셀에 저장합니다.
row_num = 1
for page_num in range(1, 9):  # 8페이지까지
    url = f"https://finance.naver.com/sise/theme.naver?&page={page_num}"
    response = requests.get(url)
    soup = BeautifulSoup(response.text, 'html.parser')

    for td_tag in soup.find_all('td', class_='col_type1'):
        theme_url = "https://finance.naver.com" + td_tag.find('a')['href']
        theme_name = td_tag.text.strip()
        stocks = get_stocks(theme_url)

        # 엑셀에 테마명과 해당 테마에 속한 주식명을 기록합니다.
        ws.cell(row=row_num, column=1, value=theme_name)
        ws.cell(row=row_num, column=2, value=', '.join(stocks))
        row_num += 1

# 결과를 엑셀 파일로 저장합니다.
wb.save("stock_data.xlsx")
